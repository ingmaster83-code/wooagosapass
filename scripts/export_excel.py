"""
gosapass 자격증 목록 엑셀 내보내기
"""
import json
import re
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"

# ── 데이터 로드 ─────────────────────────────────────────────────────────────
exams_data = json.loads((DATA_DIR / "exams.json").read_text(encoding="utf-8"))
exams = exams_data["items"]

# stats 파일에서 수수료·합격률 가져오기
def get_stats(jmcd):
    p = DATA_DIR / "stats" / f"{jmcd}.json"
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))

# 산업기사 여부 판별
def get_level(exam):
    name = exam["name"]
    series = exam.get("series", "")
    if "산업기사" in name:
        return "산업기사"
    return series

# ── 워크북 생성 ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "자격증 목록"

# 스타일 정의
HEADER_FILL   = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL  = PatternFill("solid", fgColor="DBEAFE")
SUBHEAD_FONT  = Font(bold=True, color="1E40AF", size=10)

LEVEL_COLOR = {
    "기술사":   "7C3AED",
    "기능장":   "B45309",
    "기사":     "1D4ED8",
    "산업기사": "0369A1",
    "기능사":   "15803D",
}

thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hcell(ws, row, col, value, fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border
    return c

# ── 헤더 ────────────────────────────────────────────────────────────────────
headers = [
    "No", "종목코드", "자격증명", "등급", "직무분야",
    "필기 수수료(원)", "실기 수수료(원)",
    "최근 필기 응시자", "최근 필기 합격률(%)",
    "최근 실기 응시자", "최근 실기 합격률(%)",
    "사이트 URL"
]
ws.row_dimensions[1].height = 36
for col, h in enumerate(headers, 1):
    hcell(ws, 1, col, h, fill=HEADER_FILL, font=HEADER_FONT)

# 등급별 색상 채우기 함수
def level_fill(level):
    color = LEVEL_COLOR.get(level, "6B7280")
    return PatternFill("solid", fgColor=color)

# ── 데이터 행 ────────────────────────────────────────────────────────────────
# 등급 → 정렬 순서
LEVEL_ORDER = {"기술사": 0, "기능장": 1, "기사": 2, "산업기사": 3, "기능사": 4}

sorted_exams = sorted(
    exams,
    key=lambda e: (
        LEVEL_ORDER.get(get_level(e), 9),
        e.get("field", ""),
        e["name"]
    )
)

for i, exam in enumerate(sorted_exams, 1):
    row = i + 1
    ws.row_dimensions[row].height = 20

    jmcd  = exam["jmcd"]
    name  = exam["name"]
    level = get_level(exam)
    field = exam.get("field", "")

    stats = get_stats(jmcd)
    fee   = stats.get("fee", {})
    written_stats   = stats.get("stats", {}).get("written", [])
    practical_stats = stats.get("stats", {}).get("practical", [])

    # 최근 연도 통계
    def latest(rows):
        for r in reversed(rows):
            if r.get("applicants", 0) > 0:
                return r
        return {}

    w = latest(written_stats)
    p = latest(practical_stats)

    url = f"https://wooagosapass.wooahouse.com/exam/{name}.html"

    row_data = [
        i,
        jmcd,
        name,
        level,
        field,
        int(fee["written"])   if fee.get("written")   else "",
        int(fee["practical"]) if fee.get("practical") else "",
        w.get("applicants", ""),
        w.get("passRate", ""),
        p.get("applicants", ""),
        p.get("passRate", ""),
        url,
    ]

    lf = level_fill(level)
    for col, val in enumerate(row_data, 1):
        align = "center" if col in (1, 2, 4, 5, 6, 7, 8, 9, 10, 11) else "left"
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        c.font = Font(size=10)
        if col == 4:  # 등급 열 색상
            c.fill = lf
            c.font = Font(bold=True, color="FFFFFF", size=10)
        elif row % 2 == 0:
            c.fill = PatternFill("solid", fgColor="F9FAFB")

# ── 열 너비 ──────────────────────────────────────────────────────────────────
col_widths = [6, 10, 28, 10, 16, 16, 16, 16, 16, 16, 16, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 1행 고정
ws.freeze_panes = "A2"

# ── 요약 시트 ────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("등급별 요약")
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 20

hcell(ws2, 1, 1, "등급",   fill=HEADER_FILL, font=HEADER_FONT)
hcell(ws2, 1, 2, "종목 수", fill=HEADER_FILL, font=HEADER_FONT)
hcell(ws2, 1, 3, "비고",   fill=HEADER_FILL, font=HEADER_FONT)

from collections import Counter
level_counts = Counter(get_level(e) for e in exams)
level_order  = ["기술사", "기능장", "기사", "산업기사", "기능사"]
for r, lv in enumerate(level_order, 2):
    cnt = level_counts.get(lv, 0)
    ws2.cell(row=r, column=1, value=lv).border   = border
    ws2.cell(row=r, column=2, value=cnt).border  = border
    ws2.cell(row=r, column=3, value="").border   = border
    ws2.cell(row=r, column=1).fill = level_fill(lv)
    ws2.cell(row=r, column=1).font = Font(bold=True, color="FFFFFF", size=10)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

total_row = len(level_order) + 2
ws2.cell(row=total_row, column=1, value="합계").font  = Font(bold=True, size=10)
ws2.cell(row=total_row, column=2, value=len(exams)).font = Font(bold=True, size=10)
for col in range(1, 4):
    ws2.cell(row=total_row, column=col).border    = border
    ws2.cell(row=total_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=total_row, column=col).fill      = PatternFill("solid", fgColor="F3F4F6")

# ── 저장 ────────────────────────────────────────────────────────────────────
today = date.today().strftime("%Y%m%d")
out_path = ROOT / f"gosapass_자격증목록_{today}.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
print(f"총 {len(exams)}개 자격증")
